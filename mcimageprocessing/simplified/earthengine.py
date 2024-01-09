import ee
import json
from pydantic import BaseModel, validator, root_validator
import datetime
import re
import geopandas as gpd
import requests
from typing import ClassVar
import calendar
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

import os
import geemap


class EarthEngineManager(BaseModel):
    """
    EarthEngineManager class for managing Earth Engine operations.
    """
    authentication_file: str
    year_ranges: list = []

    aggregation_functions: dict = None

    vis_params: ClassVar[dict] = {'NDVI': {'min': 0, 'max': 1,
                                           'palette': ['FFFFFF', 'CE7E45', 'DF923D', 'F1B555', 'FCD163', '99B718',
                                                       '74A901', '66A000', '529400', '3E8601', '207401', '056201',
                                                       '004C00', '023B01', '012E01', '011D01', '011301']}}
    ee_dates: list = []

    @root_validator(pre=True)
    def set_aggregation_functions(cls, values):
        """
            Set the aggregation functions for the class.

            :param values: A dictionary of aggregation functions. The keys are the names of the functions, and the values are lambda functions that take an ee.ImageCollection as input and return
        * the aggregated value.
            :type values: dict
            :return: The input values parameter.
            :rtype: dict
            """
        cls.aggregation_functions = {
            'mode': lambda ic: ic.mode(),
            'median': lambda ic: ic.median(),
            'mean': lambda ic: ic.mean(),
            'max': lambda ic: ic.max(),
            'min': lambda ic: ic.min(),
            'sum': lambda ic: ic.reduce(ee.Reducer.sum()),
            'first': lambda ic: ic.sort('system:time_start', False).first(),
            'last': lambda ic: ic.sort('system:time_start', False).last()
        }
        return values

    @root_validator(pre=True)
    def set_vis_params(cls, values):
        values["vis_params"] = {'NDVI': {'min': 0, 'max': 1,
                                         'palette': ['FFFFFF', 'CE7E45', 'DF923D', 'F1B555', 'FCD163', '99B718',
                                                     '74A901', '66A000', '529400', '3E8601', '207401', '056201',
                                                     '004C00', '023B01', '012E01', '011D01', '011301']}}
        return values

    @validator('authentication_file', pre=True)
    def load_credentials(cls, v):
        """
        Load Earth Engine credentials from an authentication file.

        :param v: The path to the authentication file.
        :type v: str
        :return: The path to the authentication file.
        :rtype: str
        """
        with open(v, "r") as f:
            ee_auth_json = f.read().strip()

        ee_auth = json.loads(ee_auth_json)

        # Initialize Earth Engine credentials
        credentials = ee.ServiceAccountCredentials(
            email=ee_auth['client_email'],
            key_data=ee_auth['private_key']
        )

        ee.Initialize(credentials)
        return v

    def helper_image_creation(self):
        """
        Helper method for image creation.

        This method prompts the user to input the image collection code and validates its format.
        If the format is invalid, a ValueError is raised.

        Next, it retrieves the image collection dates using the 'get_image_collection_dates' method.
        These dates are converted to datetime objects for further processing.

        The minimum and maximum dates from the collection are determined.

        The method then prompts the user to choose between a single date of imagery or using an aggregation function on a range of dates.
        Based on the user's choice, the available image dates are printed.

        :return: None

        """
        collection = input(
            'What is the image collection you want to analyze? Please enter the image collection code, generally in'
            'the format of \'Source/Code\' (e.g., \'COPERNICUS/S2_SR\') or \'Source/Code/Dataset\' (e.g., '
            '\'MODIS/061/MOD13A2\')"')

        pattern = r'^[A-Za-z0-9_-]+(?:/[A-Za-z0-9_-]+){1,2}$'
        if not re.match(pattern, collection):
            raise ValueError(
                f"Invalid collection format: {collection}. Expected format: 'Source/Code' (e.g., 'COPERNICUS/S2_SR') or 'Source/Code/Dataset' (e.g., 'MODIS/061/MOD13A2')")

        image_dates = self.get_image_collection_dates(collection)

        day_formated = [datetime.datetime.strptime(day, '%Y-%m-%d') for day in image_dates]

        min_date = min(day_formated)
        max_date = max(day_formated)

        aggregation_or_not = input(f'This collection contains images that range from {min_date.strftime("%Y-%m-%d")} - '
                                   f'{max_date.strftime("%Y-%m-%d")}. Are you interested in (1) a single date of imagery, or (2) using an aggregation'
                                   f'function on a range of dates? Please enter 1 or 2.')

        if aggregation_or_not == 1:
            date = input(f"Which of the following image dates do you want to retrieve? {image_dates}")

        else:
            print(f"Which of the following image dates do you want to retrieve? {image_dates}")

    @classmethod
    def validate_aggregation_function(cls, function):
        if function not in cls.aggregation_functions:
            raise ValueError(
                f"Invalid aggregation function: {function}. Must be one of: {', '.join(cls.aggregation_functions.keys())}")
        return function

    def generate_monthly_date_ranges(self, start_date, end_date):
        # Convert start and end dates from strings to date objects
        # start_date_dt = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        # end_date_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d')

        date_ranges = []

        for year in range(start_date.year, end_date.year + 1):
            for month in range(1, 13):
                # Skip months before the start month in the start year
                if year == start_date.year and month < start_date.month:
                    continue
                # Stop if the month is after the end month in the end year
                if year == end_date.year and month > end_date.month:
                    break
                # Get the last day of the month
                last_day = calendar.monthrange(year, month)[1]
                # Format the start and end dates of the month
                month_start_date = datetime.date(year, month, 1).strftime('%Y-%m-%d')
                month_end_date = datetime.date(year, month, last_day).strftime('%Y-%m-%d')
                # Append the tuple of the start and end date of the month to the list
                date_ranges.append((month_start_date, month_end_date))

        return date_ranges

    def generate_yearly_date_ranges(self, start_date, end_date):
        # Convert start and end dates from strings to date objects
        # start_date_dt = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        # end_date_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d')

        # Check if end_date is supplied and valid
        if not end_date or end_date < start_date:
            raise ValueError("The end date must be provided and must be after the start date.")

        date_ranges = []

        for year in range(start_date.year, end_date.year + 1):
            # Determine the start of the year
            if year == start_date.year:
                year_start_date = start_date
            else:
                year_start_date = datetime.date(year, 1, 1)

            # Determine the end of the year
            if year == end_date.year:
                year_end_date = end_date
            else:
                year_end_date = datetime.date(year, 12, 31)

            # Format the start and end dates of the year
            year_start_date_str = year_start_date.strftime('%Y-%m-%d')
            year_end_date_str = year_end_date.strftime('%Y-%m-%d')

            # Append the tuple of the start and end date of the year to the list
            date_ranges.append((year_start_date_str, year_end_date_str))

        return date_ranges

    def get_image_collection_dates(self, collection: str, min_max_only: bool = False):
        # Regular expression to match both formats
        pattern = r'^[A-Za-z0-9_-]+(?:/[A-Za-z0-9_-]+){1,2}$'

        # if not re.match(pattern, collection):
        #     raise ValueError(
        #         f"Invalid collection format: {collection}. Expected format: 'Source/Code' (e.g., 'COPERNICUS/S2_SR') or 'Source/Code/Dataset' (e.g., 'MODIS/061/MOD13A2')")

        def format_dates(img):
            original_date = ee.Date(img.get('system:time_start'))
            current_day = original_date.format('YYYY-MM-dd')
            return ee.Feature(None, {'current_date': current_day})

        def get_min_max_dates(collection):
            # Get the minimum date in the collection.
            min_date = ee.Date(collection.aggregate_min('system:time_start')).format('YYYY-MM-dd')
            # Get the maximum date in the collection.
            max_date = ee.Date(collection.aggregate_max('system:time_start')).format('YYYY-MM-dd')

            min_date = min_date.getInfo()

            max_date = max_date.getInfo()

            return [min_date, max_date]

        collection = ee.ImageCollection(collection)
        if min_max_only:
            return get_min_max_dates(collection)
        else:
            formatted_dates = collection.map(format_dates)
            current_date_list = formatted_dates.aggregate_array('current_date')
            ee_date_list = current_date_list.getInfo()
            self.ee_dates = [x for x in ee_date_list]
            return self.ee_dates

    # def get_boundary(self):
    #     gdf = gpd.GeoDataFrame.from_postgis(
    #         f'SELECT *, ST_SimplifyPreserveTopology(geometry, {os.getenv("POLYGON_SIMPLIFICATION_TOLERANCE")}) as simplified_geometry FROM public."state_adm"',
    #         self.sql_connect(),
    #         geom_col='simplified_geometry'
    #     )
    #     gdf.crs = "EPSG:4326"
    #
    #     comitas_states = gdf[gdf['statename'].isin(['Adamawa', 'Taraba'])]
    #     comitas_states = comitas_states.assign(dissolve_field=1)
    #
    #     dissolved_gdf = comitas_states.dissolve(by='dissolve_field')
    #
    #     geometry = dissolved_gdf.geometry.iloc[0]
    #     if not geometry.is_valid or geometry.geom_type == 'MultiPolygon':
    #         # Convert MultiPolygon to a single Polygon, here we take the largest polygon
    #         largest_polygon = max(geometry, key=lambda a: a.area)
    #         boundary_coords = list(largest_polygon.exterior.coords)
    #     else:
    #         boundary_coords = list(geometry.exterior.coords)
    #
    #     # Convert coordinates for Earth Engine
    #     ee_coords = [list(coord) for coord in boundary_coords]  # Convert to list of lists
    #
    #     return ee.Geometry.Polygon(ee_coords)

    def calculate_statistics(self, img, geometry, band):
        # Define the reducers for each statistic you want to calculate
        reducers = ee.Reducer.mean().combine(
            reducer2=ee.Reducer.sum(),
            sharedInputs=True
        ).combine(
            reducer2=ee.Reducer.max(),
            sharedInputs=True
        ).combine(
            reducer2=ee.Reducer.min(),
            sharedInputs=True
        ).combine(
            reducer2=ee.Reducer.stdDev(),
            sharedInputs=True
        ).combine(
            reducer2=ee.Reducer.variance(),
            sharedInputs=True
        ).combine(
            reducer2=ee.Reducer.median(),
            sharedInputs=True
        )

        # Apply the reducers to the image
        stats = img.reduceRegion(reducer=reducers, geometry=geometry, maxPixels=1e12)

        # Rename the keys in the dictionary to be more descriptive
        stats_dict = {
            'Mean': stats.get(band + '_mean'),
            'Sum': stats.get(band + '_sum'),
            'Max': stats.get(band + '_max'),
            'Min': stats.get(band + '_min'),
            'Standard Deviation': stats.get(band + '_stdDev'),
            'Variance': stats.get(band + '_variance'),
            'Median': stats.get(band + '_median')
        }

        return stats_dict

    def get_image(self,
                  multi_date: bool,
                  aggregation_period: str=None,
                  aggregation_method: str=None,
                  start_date:str=None,
                  end_date:str=None,
                    date:str=None,
                  scale=None,
                  image_collection:str=None,
                  band:str=None,
                  additional_filter=False,
                  filter_argument=None,
                  geometry=None,
                  statistics_only=False
                  ):
        """
        Fetches the specified image from the Earth Engine dataset and applies the specified aggregation function.

        Parameters:
        - content (dict): Dictionary containing image fetching and processing specifications.

        Returns:
        - img (ee.Image): The processed image.
        - nigeria (ee.Geometry): Geometry object representing Nigeria's boundary.
        """

        if multi_date:
            self.__class__.validate_aggregation_function(aggregation_method)

            img_function = EarthEngineManager.aggregation_functions.get(aggregation_method)
            if img_function is None:
                raise ValueError(f"Invalid aggregation function: {aggregation_method}")


            boundary = geometry

            # if additional_filter:
            #     img_collection = ee.ImageCollection(image_collection).filter(
            #         ee.Filter.date(start_date, end_date)).select(band).filter(ee.Filter.bounds(geometry))
            # else:
                # img_collection = ee.ImageCollection(image_collection).filter(
                #     ee.Filter.date(start_date, end_date)).select(band)

            img_collection = ee.ImageCollection(image_collection).filter(
                ee.Filter.date(start_date, end_date)).select(band).filter(ee.Filter.bounds(geometry))

            # if function not in self.__class__.aggregation_functions.keys():
            #     raise ValueError('Aggregation Function must be one of mode, median, mean, max, min, sum or None')


            mask = ee.Image.constant(1).clip(geometry)

            img = self.__class__.aggregation_functions[aggregation_method](img_collection).clip(boundary)

            # Step 1: Mask where value is not -9999
            valid_pixels = img.neq(-9999)

            # Step 2: Combine valid_pixels mask with nigeria_mask
            combined_mask = valid_pixels.And(mask)

            # Step 3: Use the combined mask to mask the aggregated image
            img_masked = img.updateMask(combined_mask)

            nodata_value = 0
            img = img_masked.unmask(nodata_value)


            # Checking for presence of -9999 pixels
            # min_value = img.reduceRegion(reducer=ee.Reducer.min(), geometry=nigeria, scale=30, maxPixels=1e12).get(band).getInfo()

            # if min_value == -9999:
            #     raise ValueError('The image contains -9999 pixels after processing.')

            return img, boundary
        else:

            if isinstance(geometry, ee.Feature):
                geometry = geometry.geometry()
            elif isinstance(geometry, ee.Geometry):
                geometry = geometry
            else:
                raise ValueError("The region must be a Feature or Geometry.")

            img_collection = ee.ImageCollection(image_collection).filterDate(ee.Date(date)).select(band).filter(ee.Filter.bounds(geometry))

            mask = ee.Image.constant(1).clip(geometry)

            img = img_collection.first().clip(geometry)

            # Step 1: Mask where value is not -9999
            valid_pixels = img.neq(-9999)

            # Step 2: Combine valid_pixels mask with nigeria_mask
            combined_mask = valid_pixels.And(mask)

            # Step 3: Use the combined mask to mask the aggregated image
            img_masked = img.updateMask(combined_mask)

            nodata_value = 0
            img = img_masked.unmask(nodata_value)


            # Checking for presence of -9999 pixels
            # min_value = img.reduceRegion(reducer=ee.Reducer.min(), geometry=nigeria, scale=30, maxPixels=1e12).get(band).getInfo()

            # if min_value == -9999:
            #     raise ValueError('The image contains -9999 pixels after processing.')

            return img, geometry

    def get_image_download_url(self, region, scale, img, format='GEO_TIFF', crs='EPSG:4326'):


        if isinstance(region, ee.Feature):
            geometry = region.geometry()
        elif isinstance(region, ee.Geometry):
            geometry = region
        else:
            raise ValueError("The region must be a Feature or Geometry.")


        return img.getDownloadURL({
            'region': geometry,
            'scale': scale,
            'crs': crs,
            'format': format
        })

    # def download_file_from_url_to_s3(url, bucket_name, s3_destination_key):
    #     # Initialize a session using Boto3
    #     session = boto3.session.Session()
    #     # Get the S3 resource
    #     s3 = session.resource('s3')
    #
    #     # Get the response from the URL
    #     response = requests.get(url, stream=True)
    #     response.raise_for_status()
    #
    #     # Upload the content directly to S3
    #     s3_object = s3.Object(bucket_name, s3_destination_key)
    #     s3_object.put(Body=response.raw)

    @staticmethod
    def download_file_from_url(url, destination_path):
        response = requests.get(url, stream=True)
        response.raise_for_status()

        with open(destination_path, 'wb') as file:
            for chunk in response.iter_content(chunk_size=8192):
                file.write(chunk)

    def img_min_max(self, img, scale, min_threshold=None, boundary=None, band=None):
        max_pixels = 1e12

        # If a threshold is provided, mask the image to exclude values below the threshold
        if min_threshold is not None:
            img = img.updateMask(img.gte(min_threshold))

        stats = img.reduceRegion(
            reducer=ee.Reducer.minMax(),
            bestEffort=True,
            scale=scale,
            geometry=boundary,  # Add this line
            maxPixels=max_pixels
        ).getInfo()

        min_value = stats[f"{band}_min"]
        max_value = stats[f"{band}_max"]

        return min_value, max_value

    def plot_image(self, img, vis_params):
        center_lat = 2
        center_lon = 32
        zoomlevel = 6
        map = geemap.Map(center=[center_lat, center_lon], zoom=zoomlevel)
        map.addLayer(img, vis_params=vis_params)
        map.addLayerControl()
        return map

    def process_images(self, start_date, end_date, image_collection, band, country, aggregation_type, function):
        """
        Automatically processes all images from specified start date to specified end date on a monthly/yearly aggregate.
        Starts by creating date ranges according to specified aggregation type.
        For each date range, it retrieves and processes image, then downloads the image.
        Args:
            start_date (str): Start date in YYYY-MM-DD format.
            end_date (str): End date in YYYY-MM-DD format.
            image_collection (str): Image collection code.
            band (str): The band of the image collection to select.
            aggregation_type (str): Monthly or Yearly.

        Returns:
            None
        """
        # Validate aggregation type
        if aggregation_type.lower() not in ("monthly", "yearly"):
            raise ValueError('Aggregation type should be either "monthly" or "yearly".')

        # Generate date ranges
        if aggregation_type.lower() == "monthly":
            date_ranges = self.generate_monthly_date_ranges(int(start_date[:4]))
        else:
            date_ranges = self.generate_yearly_date_ranges(int(start_date[:4]))

        for start, end in date_ranges:
            start = datetime.datetime.strptime(start, "%Y-%m-%d").date()
            end = datetime.datetime.strptime(end, "%Y-%m-%d").date()

            # If the date range is within the specified start date and end date
            if start >= datetime.datetime.strptime(start_date, "%Y-%m-%d").date() and end <= datetime.datetime.strptime(end_date,
                                                                                                      "%Y-%m-%d").date():
                print(f"Processing images from {start} to {end}...")

                # Proceed to fetch and process image
                img, boundary = self.get_image(function, str(start), str(end), image_collection, band, country=country)

                # Generate download URL
                download_url = self.get_image_download_url(boundary, 1000, img)

                # Prompt the download URL
                print("Download URL: ", download_url)

                # Download the image
                destination_path = f"{band}_{start}_{end}.tif"
                self.download_file_from_url(download_url, destination_path)

    def get_admin_units(self, level=0):
        # Load the dataset
        gaul_dataset = ee.FeatureCollection(f"FAO/GAUL_SIMPLIFIED_500m/2015/level{level}")

        if level == 0:
            unique_countries = gaul_dataset.aggregate_array('ADM0_NAME').distinct()
            return sorted(unique_countries.getInfo())

        elif level == 1 or level == 2:
            # Get the distinct higher-level units
            unique_higher_level_units = gaul_dataset.aggregate_array(f'ADM{level - 1}_NAME').distinct()

            # Create a FeatureCollection where each feature is a higher-level unit and its property is the list of lower-level units
            def process_higher_level_unit(current, prev):
                prev = ee.Dictionary(prev)
                current = ee.String(current)

                filtered_1 = gaul_dataset.filter(ee.Filter.eq(f'ADM{level - 1}_NAME', current))
                lower_level_units_1 = filtered_1.aggregate_array(f'ADM{level}_NAME').distinct()

                # if level == 2:
                #     output_dict = ee.Dictionary({})
                #
                #     def process_higher_level_unit_2(level_1_unit, prev_2):
                #         prev_2 = ee.Dictionary(prev_2)
                #         level_1_unit = ee.String(level_1_unit)
                #
                #         # filter level 2 units for current level 1 unit
                #         filtered_2 = gaul_dataset.filter(ee.Filter.eq(f'ADM1_NAME', level_1_unit))
                #         lower_level_units_2 = filtered_2.aggregate_array('ADM2_NAME').distinct()
                #
                #         return prev_2.set(level_1_unit, lower_level_units_2)
                #
                #     lower_level_units_2_dict = ee.Dictionary(
                #         ee.List(lower_level_units_1.iterate(process_higher_level_unit_2, ee.Dictionary({}))))
                #     return prev.set(current, lower_level_units_2_dict)
                #
                # else:
                return prev.set(current, lower_level_units_1)

            admin_units_dict = ee.Dictionary(
                ee.List(unique_higher_level_units.iterate(process_higher_level_unit, ee.Dictionary({}))))

            return admin_units_dict.getInfo()

    def create_excel(self, datatype='EE', indicator=None, admin_level=0, destination=None):
        admin_levels = self.get_admin_units(level=admin_level)

        # Define the structure of the workbook sheets
        sheet_names = {
            'PrimaryFilters': {
                'colnames': ['Admin0', 'StartDate', 'EndDate', 'AggregationMethod', 'Frequency', 'Admin1'],
                'colletters': ['A', 'B', 'C', 'D', 'E', 'F']
            },
            'Options': {
                'colnames': ['Admin0 Options'],
                'colletters': ['A'],
                'list_location': {
                    'Admin0': admin_levels
                }
            }
        }

        wb = Workbook()

        # Create sheets and set column headers
        for sheet, properties in sheet_names.items():
            wb.create_sheet(sheet)
            current_sheet = wb[sheet]
            for col_name, col_letter in zip(properties['colnames'], properties['colletters']):
                current_sheet[f"{col_letter}1"] = col_name

        # Populate the 'Options' sheet
        options_sheet = wb['Options']
        for idx, option in enumerate(admin_levels, start=2):  # Start at row 2
            options_sheet[f"A{idx}"] = option

        # Add data validation to 'PrimaryFilters'
        primary_filters_sheet = wb['PrimaryFilters']

        dv = DataValidation(type="list", formula1=f"Options!$A$2:$A${len(admin_levels) + 1}",
                            showDropDown=False, showInputMessage=True, errorStyle="stop", showErrorMessage=True)

        for row in range(2, 500):  # Assuming you want to apply this to all possible rows
            dv.add(primary_filters_sheet[f"{get_column_letter(1)}{row}"])

        primary_filters_sheet.add_data_validation(dv)

        # if admin_level > 0:
        #     hidden_sheet = wb.create_sheet(title="Hidden")
        #     hidden_sheet.sheet_state = 'hidden'
        #
        #     for row_index, (admin0, admin1_list) in enumerate(admin_levels.items(), start=1):
        #         for col_index, admin1 in enumerate(admin1_list, start=1):
        #             hidden_sheet.cell(row=row_index, column=col_index, value=admin1)
        #         # Corrected way to create a named range
        #         named_range = f"'{hidden_sheet.title}'!$A${row_index}:$A${row_index}"
        #         defined_name = DefinedName(name=admin0, attr_text=named_range)
        #         wb.defined_names.append(defined_name)
        #
        #     # INDIRECT data validation for Admin1
        #     dv_admin1 = DataValidation(type="list", formula1=f'INDIRECT(${get_column_letter(1)}2)',
        #                                showDropDown=False)
        #     primary_filters_sheet.add_data_validation(dv_admin1)
        #     for row in range(2, 501):  # Apply to rows 2 to 500
        #         dv_admin1.add(primary_filters_sheet[f'F{row}'])

        dv_date = DataValidation(type="date",
                                 operator="between",
                                 formula1=min(self.ee_dates),
                                 formula2=max(self.ee_dates),
                                 showInputMessage=True,
                                 showErrorMessage=True,
                                 errorTitle="Invalid input",
                                 error="Enter a date between {} and {}".format(min(self.ee_dates), max(self.ee_dates)),
                                 promptTitle="Enter Date",
                                 prompt=f"Enter a date in format: YYYY-MM-DD. Date must be between {min(self.ee_dates)} and {max(self.ee_dates)}")

        # Apply to each cell in the relevant columns ('StartDate'-'B', 'EndDate'-'C')
        for col in ['B', 'C']:
            dv_date.add(f"{col}2:{col}500")  # Apply to a range instead of individual cells

        primary_filters_sheet.add_data_validation(dv_date)

        # Data validation for Aggregation Method column
        agg_methods = ['None', 'mode', 'median', 'mean', 'max', 'min', 'sum', 'first', 'last']
        dv_agg = DataValidation(type="list", formula1='"{}"'.format(','.join(agg_methods)),
                                showDropDown=False, showInputMessage=True, errorStyle="stop", showErrorMessage=True)
        dv_agg.add(f"D2:D500")  # Assuming 'AggregationMethod' is in column D
        primary_filters_sheet.add_data_validation(dv_agg)

        # Data validation for Frequency column
        frequencies = ['None', 'Monthly', 'Yearly']
        dv_freq = DataValidation(type="list", formula1='"{}"'.format(','.join(frequencies)),
                                 showDropDown=False, showInputMessage=True, errorStyle="stop", showErrorMessage=True)
        dv_freq.add(f"E2:E500")  # Assuming 'Frequency' is in column E
        primary_filters_sheet.add_data_validation(dv_freq)

        # Remove the default 'Sheet' created by openpyxl
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        return wb


# class GloFasManager(BaseModel):
#     data_dicts = {
#         "products": {
#             'cems-glofas-seasonal': {
#                 "system_version": ['operational', 'version_3_1', 'version_2_2'],
#                 'hydrological_model': ['lisflood', 'HTESSEL-LISFLOOD'],
#                 "variable": "river_discharge_in_the_last_24_hours",
#                 "leadtime_hour": [
#                     24, 48, 72, 96, 120, 144, 168, 192, 216, 240, 264, 288, 312, 336, 360, 384, 408,
#                     432, 456, 480, 504, 528, 552, 576, 600, 624, 648, 672, 696, 720, 744, 768, 792,
#                     816, 840, 864, 888, 912, 936, 960, 984, 1008, 1032, 1056, 1080, 1104, 1128, 1152,
#                     1176, 1200, 1224, 1248, 1272, 1296, 1320, 1344, 1368, 1392, 1416, 1440, 1464, 1488,
#                     1512, 1536, 1560, 1584, 1608, 1632, 1656, 1680, 1704, 1728, 1752, 1776, 1800, 1824,
#                     1848, 1872, 1896, 1920, 1944, 1968, 1992, 2016, 2040, 2064, 2088, 2112, 2136, 2160,
#                     2184, 2208, 2232, 2256, 2280, 2304, 2328, 2352, 2376, 2400, 2424, 2448, 2472, 2496,
#                     2520, 2544, 2568, 2592, 2616, 2640, 2664, 2688, 2712, 2736, 2760, 2784, 2808, 2832,
#                     2856, 2880, 2904, 2928, 2952, 2976, 3000, 3024, 3048, 3072, 3096, 3120, 3144, 3168,
#                     3192, 3216, 3240, 3264, 3288, 3312, 3336, 3360, 3384, 3408, 3432, 3456, 3480, 3504,
#                     3528, 3552, 3576, 3600, 3624, 3648, 3672, 3696, 3720, 3744, 3768, 3792, 3816, 3840,
#                     3864, 3888, 3912, 3936, 3960, 3984, 4008, 4032, 4056, 4080, 4104, 4128, 4152, 4176,
#                     4200, 4224, 4248, 4272, 4296, 4320, 4344, 4368, 4392, 4416, 4440, 4464, 4488, 4512,
#                     4536, 4560, 4584, 4608, 4632, 4656, 4680, 4704, 4728, 4752, 4776, 4800, 4824, 4848,
#                     4872, 4896, 4920, 4944, 4968, 4992, 5016, 5040, 5064, 5088, 5112, 5136, 5160],
#                 "year": [2020, 2021, 2022, 2023],
#                 "month": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
#                           "11", "12"],
#                 # "area": [10.95, -90.95, -30.95, -29.95],
#                 "format": "grib"
#             },
#             'cems-glofas-forecast': {
#                 "system_version": ['operational', 'version_3_1', 'version_2_1'],
#                 'hydrological_model': ['lisflood', 'htessel_lisflood'],
#                 'product_type': [
#                     'control_forecast', 'ensemble_perturbed_forecasts',
#                 ],
#                 "variable": "river_discharge_in_the_last_24_hours",
#                 "leadtime_hour": [
#                     24, 48, 72, 96, 120, 144, 168, 192, 216, 240, 264, 288, 312, 336, 360, 384, 408,
#                     432, 456, 480, 504, 528, 552, 576, 600, 624, 648, 672, 696, 720],
#                 "year": [2019, 2020, 2021, 2022, 2023],
#                 "month": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
#                           "11", "12"],
#                 "day": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12,
#                         13, 14, 15, 16, 17, 18, 19, 20, 21, 22,
#                         23, 24, 25, 26, 27, 28, 29, 30, 31],
#                 # "area": [10.95, -90.95, -30.95, -29.95],
#                 "format": "grib"
#             },
#             'cems-glofas-reforecast': {
#                 "system_version": ['version_4_0', 'version_3_1', 'version_2_2'],
#                 'hydrological_model': ['lisflood', 'htessel_lisflood'],
#                 'product_type': [
#                     'control_forecast', 'ensemble_perturbed_forecasts',
#                 ],
#                 "leadtime_hour": [
#                     24, 48, 72, 96, 120, 144, 168, 192, 216, 240, 264, 288, 312, 336, 360, 384, 408,
#                     432, 456, 480, 504, 528, 552, 576, 600, 624, 648, 672, 696, 720, 744, 768, 792,
#                     816, 840, 864, 888, 912, 936, 960, 984, 1008, 1032, 1056, 1080, 1104],
#                 "year": [1999, 2000, 20001, 2002, 2003, 2004, 2005, 2006, 2007,
#                          2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015,
#                          2016, 2017, 2018, 2019, 2020, 2021, 2022],
#                 "month": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
#                           "11", "12"],
#                 "day": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12,
#                         13, 14, 15, 16, 17, 18, 19, 20, 21, 22,
#                         23, 24, 25, 26, 27, 28, 29, 30, 31],
#                 # "area": [10.95, -90.95, -30.95, -29.95],
#                 "format": "grib"
#             },
#             'cems-glofas-seasonal-reforecast': {
#                 "system_version": ['version_4_0', 'version_3_1', 'version_2_2'],
#                 'hydrological_model': ['lisflood', 'htessel_lisflood'],
#                 "variable": "river_discharge_in_the_last_24_hours",
#                 "leadtime_hour": [
#                     24, 48, 72, 96, 120, 144, 168, 192, 216, 240, 264, 288, 312, 336, 360, 384, 408,
#                     432, 456, 480, 504, 528, 552, 576, 600, 624, 648, 672, 696, 720, 744, 768, 792,
#                     816, 840, 864, 888, 912, 936, 960, 984, 1008, 1032, 1056, 1080, 1104,
#                     1128, 1152, 1176, 1200, 1224, 1248, 1272, 1296, 1320, 1344, 1368, 1392, 1416, 1440, 1464, 1488,
#                     1512, 1536, 1560, 1584, 1608, 1632, 1656, 1680, 1704, 1728, 1752, 1776, 1800, 1824,
#                     1848, 1872, 1896, 1920, 1944, 1968, 1992, 2016, 2040, 2064, 2088, 2112, 2136, 2160,
#                     2184, 2208, 2232, 2256, 2280, 2304, 2328, 2352, 2376, 2400, 2424, 2448, 2472, 2496,
#                     2520, 2544, 2568, 2592, 2616, 2640, 2664, 2688, 2712, 2736, 2760, 2784, 2808, 2832,
#                     2856, 2880, 2904, 2928, 2952, 2976, 3000, 3024, 3048, 3072, 3096, 3120, 3144, 3168,
#                     3192, 3216, 3240, 3264, 3288, 3312, 3336, 3360, 3384, 3408, 3432, 3456, 3480, 3504,
#                     3528, 3552, 3576, 3600, 3624, 3648, 3672, 3696, 3720, 3744, 3768, 3792, 3816, 3840,
#                     3864, 3888, 3912, 3936, 3960, 3984, 4008, 4032, 4056, 4080, 4104, 4128, 4152, 4176,
#                     4200, 4224, 4248, 4272, 4296, 4320, 4344, 4368, 4392, 4416, 4440, 4464, 4488, 4512,
#                     4536, 4560, 4584, 4608, 4632, 4656, 4680, 4704, 4728, 4752, 4776, 4800, 4824, 4848,
#                     4872, 4896, 4920, 4944, 4968, 4992, 5016, 5040, 5064, 5088, 5112, 5136, 5160],
#
#                 "year": [1981, 1982, 1983, 1984, 1985, 1986, 1987, 1988, 1989, 1990,
#                          1991, 1992, 1993, 1994, 1995, 1996, 1997, 1998, 1999, 2000,
#                          20001, 2002, 2003, 2004, 2005, 2006, 2007,
#                          2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015,
#                          2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023],
#                 "month": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
#                           "11", "12"],
#                 "format": "grib"
#             },
#             'cems-glofas-historical': {
#                 "system_version": ['version_4_0', 'version_3_1', 'version_2_1'],
#                 'hydrological_model': ['lisflood', 'htessel_lisflood'],
#                 "product_type": ['consoloated', 'intermediate']
#                 "variable": "river_discharge_in_the_last_24_hours",
#                 "year": [1979, 1980, 1981, 1982, 1983, 1984, 1985, 1986,
#                          1987, 1988, 1989, 1990, 1991, 1992, 1993, 1994,
#                          1995, 1996, 1997, 1998, 1999, 2000, 20001, 2002,
#                          2003, 2004, 2005, 2006, 2007, 2008, 2009, 2010,
#                          2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018,
#                          2019, 2020, 2021, 2022, 2023],
#                 "month": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
#                           "11", "12"],
#                 "day": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11,
#                         12, 13, 14, 15, 16, 17, 18, 19, 20, 21,
#                         22, 23, 24, 25, 26, 27, 28, 29, 30, 31],
#             }
#         }
#     }
